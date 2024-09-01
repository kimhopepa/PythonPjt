import os.path
import json

import pandas as pd
from lib.libLog import Logger
# from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
# from openpyxl.styles import Alignment, PatternFill
import pandas as pd
# from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
import re
import chardet

import inspect  # 현재 실행되고 있는 함수 이름 체크

# from lib import libConfig
# from enum import Enum
# import traceback
# import openpyxl

# config_handler = libConfig.ConfigHandler('config.ini')
# Logger.logger = libLog.Logger.logger(config_level= config_handler.config_dict["system"]["log_level"])

COL_FILE_NAME = "파일 이름"
COL_FILE_PATH = "파일 경로"

COL_CR_CLASS = '분류'
COL_CR_ITEM = '코드 리뷰 항목'
COL_CR_RESULT = '코드 리뷰 결과'
COL_CR_RESULT_CODE = 'Code'
COL_CR_LINE = 'Line'
COL_CR_RESULT_DETAIL = '상세 내용'
COL_CR_YN = 'Y/N'

ROW_CR_RESULT_OK = 'OK'
ROW_CR_RESULT_NG = 'NG'
ROW_CR_RESULT_NONE = 'N/A'

ROW_CR_CHECK_ALL = 'ALL'
ROW_CR_CHECK_SVR = 'SVR'
ROW_CR_CHECK_CLI = 'CLI'
ROW_CR_CHECK_NONE = 'NONE'

# class COL(Enum) :
#     FileName = 1
#     FilePath = 2asdf

ROW_CR_CLASS_PERFORMANCE = '성능'
ROW_CR_CLASS_DB = 'DB'
ROW_CR_CLASS_QUERY = 'Query 검증'
ROW_CR_CLASS_STANDARD = '코드 표준'

# [성능]
# COL_CR_CLASS(분류) | COL_CR_ITEM(코드 리뷰 항목) | COL_CR_RESULT(결과) | COL_CR_LINE(Line) | COL_CR_RESULT_DETAIL(상세 내용) | SVR/CLI
CR_ITEM_IDX = 1
ROW_CR_ITEM_ACTIVE = [ROW_CR_CLASS_PERFORMANCE, '서버 스크립트 Active 감시', '', '','', ROW_CR_RESULT_NONE, ROW_CR_CHECK_SVR]
ROW_CR_ITEM_LOOP = [ROW_CR_CLASS_PERFORMANCE, 'Loop문내 처리 조건 확인',  '', '','', ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
ROW_CR_ITEM_EVENT_CHANGE = [ROW_CR_CLASS_PERFORMANCE, 'Event 교환 횟수 최소화', '', '','', ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
ROW_CR_ITEM_PROPER_DP_FCT = [ROW_CR_CLASS_PERFORMANCE, '적절한 DP 처리 함수','', '','', ROW_CR_RESULT_NONE, ROW_CR_CHECK_NONE]
ROW_CR_ITEM_DP_QUERY_OPT = [ROW_CR_CLASS_PERFORMANCE, 'DP Query 최적화 구현', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_NONE]
ROW_CR_ITEM_RAIMA_UP = [ROW_CR_CLASS_PERFORMANCE, 'RAIMA DB 증가','', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]

# [DB]
# ROW_CR_ITEM_DB_BIND = [ROW_CR_CLASS_DB, 'Query 정합성 검증 확인', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
# ROW_CR_ITEM_DB_BIND = [ROW_CR_CLASS_DB, 'DB Query 바인딩 처리', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
# ROW_CR_ITEM_DB_COMMENT = [ROW_CR_CLASS_DB, 'Query 주석 작성', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
# ROW_CR_ITEM_DB_EXCEPTION = [ROW_CR_CLASS_DB, 'DB Query 예외 처리', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
ROW_CR_ITEM_QUERY = [ROW_CR_CLASS_QUERY, 'Query 정합성 검증 확인', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_NONE]

# [코드 표준]
ROW_CR_ITEM_DP_EXCEPTION = [ROW_CR_CLASS_STANDARD, 'DP 함수 예외 처리', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_SVR]
ROW_CR_ITEM_TRY_EXCEPTION = [ROW_CR_CLASS_STANDARD, 'Try/Catch 예외 처리', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
ROW_CR_ITEM_VERSION = [ROW_CR_CLASS_STANDARD, '이력 정보 작성 확인', '', '','', ROW_CR_RESULT_NONE, ROW_CR_CHECK_SVR]
ROW_CR_ITEM_CONSTRAINTS = [ROW_CR_CLASS_STANDARD, '예상 못한 Logic 동작 확인', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_NONE]
ROW_CR_ITEM_HARD_CODE = [ROW_CR_CLASS_STANDARD, '하드 코딩 금지', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]
ROW_CR_ITEM_UNNECESSARY_CODE = [ROW_CR_CLASS_STANDARD, '불필요한 코드 금지', '', '', '',ROW_CR_RESULT_NONE, ROW_CR_CHECK_ALL]

class CodeReviewCheck:

    function_body_list =[]
    text_list = []
    # df_crc_info 컬럼 명 : 파일 이름 | 파일 Full Path
    df_crc_info = pd.DataFrame(columns=[COL_FILE_NAME, COL_FILE_PATH])

    # df_crc_result 컬럼 명 : 구분 | 코드 리뷰 항목 | 코드 리뷰 결과 | Code | Line | 상세 내용 | YN
    # df_crc_result 컬럼을 변경할 경우 ROW_CR_ITEM_ACTIVE~ ROW_CR_ITEM_UNNECESSARY_CODE 초기화 갑 수정 필요
    df_crc_result = pd.DataFrame(
        columns=[COL_CR_CLASS, COL_CR_ITEM, COL_CR_RESULT_DETAIL , COL_CR_RESULT_CODE, COL_CR_LINE, COL_CR_RESULT, COL_CR_YN])

    # Data Table 처리 동작
    class CodeData:
        # Data Table 초기화 -> df_crc_result
        @staticmethod
        def init_check_list():
            try:
                Logger.debug("CodeReviewCheck.init_check_list Call")

                #1. Data 모두 삭제
                CodeReviewCheck.df_crc_result.drop(CodeReviewCheck.df_crc_result.index, inplace=True)

                #2. Data 기본 데이터로 저장
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_QUERY)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_ACTIVE)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_LOOP)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_EVENT_CHANGE)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_PROPER_DP_FCT)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_RAIMA_UP)
                # CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_DP_QUERY_OPT)
                # CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_DB_BIND)
                # CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_DB_COMMENT)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_DP_EXCEPTION)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_TRY_EXCEPTION)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_VERSION)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_CONSTRAINTS)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_HARD_CODE)
                CodeReviewCheck.df_crc_result = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_result, ROW_CR_ITEM_UNNECESSARY_CODE)

            except Exception as e:
                Logger.error("CodeReviewCheck.init_check_list - Exception : " + str(e))

        # 선택한 파일 리스트를 df 저장 -> df_crc_info ("파일 이름", "파일 전체 경로")
        @staticmethod
        def init_file_list(selected_files):
            try:
                Logger.debug("CodeReviewCheck.init_file_list Start")
                CodeReviewCheck.df_crc_info.drop(CodeReviewCheck.df_crc_info.index, inplace=True)

                for index, file_name in enumerate(selected_files):
                    base_name = os.path.basename(file_name)
                    CodeReviewCheck.df_crc_info = CodeReviewCheck.CodeData.df_concat(CodeReviewCheck.df_crc_info, {COL_FILE_NAME: base_name, COL_FILE_PATH: file_name})


            except Exception as e:
                Logger.error("CodeReviewCheck.init_file_list - Exception : " + str(e))

        # Data Table에서 특정 컬럼 데이터만 가져오기 -> df_crc_result ("분류", "코드 리뷰 항목", 코드 리뷰 결과")
        @staticmethod
        def get_tablewidget_df():
            try:
                Logger.debug("CodeReviewCheck.get_tablewidget_df Start")

                df_item_unique = CodeReviewCheck.df_crc_result.drop_duplicates(subset = COL_CR_ITEM, keep = 'first')[[COL_CR_CLASS, COL_CR_ITEM, COL_CR_RESULT]]
                return df_item_unique
                # return CodeReviewCheck.df_crc_result[[COL_CR_CLASS, COL_CR_ITEM, COL_CR_RESULT]]
            except Exception as e:
                Logger.error("CodeReviewCheck.get_tablewidget_df - Exception : " + str(e))

        # Data Table에서 특정 컬럼 데이터만 가져오기 + 조건("코드 리뷰 항목" == cr_item_name) -> df_crc_result ("코드 리뷰 항목","상세 내용", "코드 리뷰 결과", "위치")
        @staticmethod
        def get_tablewidget_detail_df(cr_item_name):
            try:
                Logger.debug("CodeReviewCheck.get_tablewidget_detail_df Start")
                # dt_data에 index 컬럼 추가
                # 상세 내용 컬럼 순서 : 코드 리뷰 항목(COL_CR_ITEM) | 상세 내용(COL_CR_RESULT_DETAIL), Line(COL_CR_LINE),  코드 리뷰 결과(COL_CR_RESULT)
                dt_data = CodeReviewCheck.df_crc_result.loc[CodeReviewCheck.df_crc_result[COL_CR_ITEM] == cr_item_name,
                                                         [COL_CR_ITEM, COL_CR_RESULT_DETAIL, COL_CR_LINE, COL_CR_RESULT]]
                dt_data = dt_data.reset_index(drop=True)
                dt_data.loc[:, "No"] = dt_data.index + 1
                cols = ["No"] + [col for col in dt_data.columns if col != 'No']
                dt_data = dt_data[cols]

                return dt_data

            except Exception as e:
                Logger.error("CodeReviewCheck.get_tablewidget_detail_df - Exception : " + str(e))

        @staticmethod
        def get_dict_from_list(input_list, key="selectedList"):
            try:
                Logger.debug("CodeReviewCheck.get_dict_from_list Start")
                result_dict = {}

                for item in input_list:
                    result_dict[key] = item

                return result_dict
            except Exception as e:
                Logger.error("CodeReviewCheck.get_dict_from_list - Exception : " + str(e))

        # DataFrame에서 Dictionary를 추가 하여 반환
        @classmethod
        def df_concat(cls, update_df, new_list):
            try:
                Logger.debug("CodeReviewCheck - df_concat Call" + str(new_list))
                # new_record_df = pd.DataFrame([new_dict])
                # update_df = pd.concat([update_df, new_record_df], ignore_index=True)
                update_df.loc[len(update_df)] = new_list

            except Exception as e:
                Logger.error("CodeReviewCheck.df_concat - Exception : " + str(e))
            return update_df

        # 파일 이름에서 확장자를 제거하여 반환
        @staticmethod
        def get_remove_extension(filename):
            try:
                Logger.debug("CodeReviewCheck - get_remove_extension start")
                name, ext = os.path.splitext(filename)
                return name

            except Exception as e:
                Logger.error("CodeReviewCheck.get_remove_extension - Exception : " + str(e))

        @classmethod
        def get_file_path(cls, file_name):
            file_path = None
            try:
                Logger.debug("CodeReviewCheck - get_file_path start")
                file_path = CodeReviewCheck.df_crc_info[CodeReviewCheck.df_crc_info[COL_FILE_NAME] == file_name][COL_FILE_PATH].iloc[0]
                return file_path

            except Exception as e:
                Logger.error("CodeReviewCheck.get_file_path - Exception : " + str(e))


    # UI 관련 동작 구현
    class CodeUI:
        '''
        function : df_concat
        args : cls(클래스 인스턴스), update_df(변경될 DataFrame), new_dict(DataFrame에 추가`````````````````````````````````````````될 Dictionary)
        detail : Dicionary 데이터를 DataFrame 변수에 저장하여 반환
        return : 변경된 DataFrame을 반환
        '''

        # 테이블 Widget 데이터 -> DafaFrame으로 변환
        @staticmethod
        def get_df_to_tablewidget(tb_widget):
            try:
                Logger.debug("CodeReviewCheck.file_export Start")
                # 1. Table_Widget -> DataFrame으로 변환
                # 열 헤더 가져오기
                headers = [tb_widget.horizontalHeaderItem(i).text() for i in range(tb_widget.columnCount())]

                # 데이터 가져오기
                data = []
                for row in range(tb_widget.rowCount()):
                    row_data = []
                    for column in range(tb_widget.columnCount()):
                        item = tb_widget.item(row, column)
                        row_data.append(item.text() if item else "")
                    data.append(row_data)

                # DataFrame으로 변환
                df = pd.DataFrame(data, columns=headers)
                return df

            except Exception as e:
                Logger.error("CodeReviewCheck.file_export - Exception : " + str(e))

        # 엑셀 파일저장
        @staticmethod
        def excel_save(save_path, df_table):
            try:
                Logger.debug("CodeReviewCheck - excel_save start")
                # 1. 엑셀 워크북 생성
                wb = Workbook()
                ws = wb.active

                # 2. DataFrame 데이터 워크북 저장
                for r in dataframe_to_rows(df_table, index=False, header=True):
                    ws.append(r)

                # 3. 모든 셀 가운데 정렬 및 맞춤
                # cell 배경
                gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

                # 윤곽선 스타일 정의
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                CodeReviewCheck.CodeUI.merge_cells_by_column(ws, 1)
                CodeReviewCheck.CodeUI.merge_cells_by_column(ws, 2)

                # 셀 정렬 및 스타일 적용
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.row == 1:
                            cell.fill = gray_fill
                        if(cell.column) == ws.max_column :
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                        # 모든 셀에 윤곽선 추가
                        cell.border = thin_border

                # 각 열의 너비를 셀 내용에 맞게 자동 조정
                for column_cells in ws.columns:
                    max_length = 0
                    column = column_cells[0].column_letter  # 열의 문자 (예: 'A', 'B', 'C')
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    adjusted_width = (max_length * 2 + 2)
                    ws.column_dimensions[column].width = adjusted_width

                # CodeReviewCheck.CodeUI.excel_merge_cell(ws, 'A2', 'A7')
                # CodeReviewCheck.CodeUI.excel_merge_cell(ws, 'A8', 'A9')
                # CodeReviewCheck.CodeUI.excel_merge_cell(ws, 'A10', 'A15')
                #
                # CodeReviewCheck.CodeUI.set_column_width(ws, 1, 20)
                # CodeReviewCheck.CodeUI.set_column_width(ws, 2, 40)
                # CodeReviewCheck.CodeUI.set_column_width(ws, 3, 30)

                # 3-1 Cell 병합
                # 병합할 셀 영역
                # merge_range = 'A2:A7'
                #
                # # 병합된 셀의 값을 첫 번째 셀의 값으로 설정
                # first_cell = ws['A2']
                # value_to_merge = first_cell.value
                # ws.merge_cells(merge_range)
                #
                # # 병합된 셀에 값 설정
                # merged_cell = ws['A2']
                # merged_cell.value = value_to_merge

                # 4. 엑셀 파일 저장
                wb.save(save_path)
            except Exception as e:
                Logger.error("CodeReviewCheck.excel_save - Exception : " + str(e))
                # print(traceback.format_exc())

        @classmethod
        def merge_cells_by_column(cls, worksheet, col_idx):
            try:
                start_row = 2  # 데이터 시작 행 (헤더 제외)
                end_row = worksheet.max_row
                current_value = None
                merge_start = start_row

                for row in range(start_row, end_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx)
                    if cell.value != current_value:
                        if current_value is not None:
                            worksheet.merge_cells(start_row=merge_start, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                            worksheet.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical='center')
                        current_value = cell.value
                        merge_start = row

                # 마지막 셀 병합
                worksheet.merge_cells(start_row=merge_start, start_column=col_idx, end_row=end_row, end_column=col_idx)
                worksheet.cell(row=merge_start, column=col_idx).alignment = Alignment(vertical='center')

            except Exception as e:
                Logger.error("CodeReviewCheck.merge_cells_by_column - Exception : " + str(e))

        # 엑셀 시트 Cell 병합
        @classmethod
        def excel_merge_cell(cls, ws, start_cell, end_cell):
            try:
                Logger.debug("CodeReviewCheck - excel_merge_cell start")
                # 병합할 셀 영역
                merge_range = f'{start_cell}:{end_cell}'

                # 병합된 셀의 값을 첫 번째 셀의 값으로 설정
                first_cell = ws[start_cell]
                value_to_merge = first_cell.value
                ws.merge_cells(merge_range)

                # 병합된 셀에 값 설정
                merged_cell = ws[start_cell]
                merged_cell.value = value_to_merge
            except Exception as e:
                Logger.error("CodeReviewCheck.excel_merge_cell - Exception : " + str(e))
                # print(traceback.format_exc())

        # 엑셀 시트 열 사이즈 설정
        @classmethod
        def set_column_width(cls, sheet, column, width):
            try:
                Logger.debug("CodeReviewCheck - set_column_width start")
                col_letter = get_column_letter(column)
                sheet.column_dimensions[col_letter].width = width
            except Exception as e:
                Logger.error("CodeReviewCheck.set_column_width - Exception : " + str(e))

        @classmethod
        def get_export_df(cls):
            try :
                Logger.info("CodeReviewCheck.get_export_df Call")

                # Export할 컬럼을 설정 : '분류', '코드 리뷰 항목', '코드 리뷰 결과', 'Code' ,'Line'
                export_df = CodeReviewCheck.df_crc_result[
                    [COL_CR_CLASS, COL_CR_ITEM, COL_CR_RESULT, COL_CR_LINE, COL_CR_RESULT_CODE, COL_CR_RESULT_DETAIL]]
                return export_df
            except Exception as e:
                Logger.error("CodeReviewCheck.set_column_width - Exception : " + str(e))


    # 코드 리뷰 검증 기능 구현 클래스
    class CodeCheck:

        # region 1.Code parsing을 위한 함수 모음
        @classmethod     # 주석으로 공백으로 제거
        def remove_line_comments(cls, code:str) -> str:
            try:
                lines = code.splitlines()
                modified_lines = []
                for line in lines:
                    # Remove comments starting with /
                    if '/' in line:
                        line = line.split('/')[0] + ' ' * len(line.split('//')[1])
                    modified_lines.append(line)
                return '\n'.join(modified_lines)
            except Exception as e:
                Logger.error("CodeReviewCheck.remove_line_comments - Exception : " + str(e))

        # text의 시작 위치와 마지막 위치를 공백으로 변경 -> 필요 없는 코드를 삭제하는 경우 사용
        @classmethod
        def replace_with_spaces(cls, text: str, start_pos: int, end_pos: int) -> str:
            try :
                # 인덱스의 유효성 검사
                if start_pos < 0:
                    start = 0
                if end_pos > len(text):
                    end = len(text)
                if start_pos >= end_pos:
                    return text

                # 현재 공백으로 만들 텍스트 길이 체크
                length_to_replace = end_pos - start_pos

                # 시작 index와 마지막 index사이의 코드를 공백으로 만듬 (체크시 필요 없는 코드)
                replace_str = text[:start_pos] + ' ' * length_to_replace + text[end_pos:]
                return replace_str

            except Exception as e:
                Logger.error("CodeReviewCheck.replace_with_spaces - Exception : " + str(e))

        @classmethod
        def is_check_pattern(cls, text: str, pattern : str) -> bool:
            try :
                # 정규식 패턴 정의 (주석 없이 delay*); 패턴을 찾음)
                pattern = r'delay.*\);'

                # 정규식 검색
                match = re.search(pattern, text)
                return match is not None
            except Exception as e :
                Logger.error("CodeReviewCheck.is_check_pattern - Exception : " + str(e))

        @classmethod
        def find_closing_brace_index(cls, text, open_brace_index):
            try:
                stack = 0
                for index, char in enumerate(text[open_brace_index:], start=open_brace_index):
                    if char == '{':
                        stack += 1
                    elif char == '}':
                        stack -= 1
                        if stack == 0:
                            return index
                return -1
            except Exception as e:
                Logger.error("CodeReviewCheck.find_closing_brace_index - Exception : " + str(e))
        # endregion

        # region 2. DataTable, 파일 등 공통으로 사용하는 함수 모음
        # 파일을 읽어서 텍스트 반환
        @staticmethod
        def get_file_to_text(file_path):
            try:
                with open(file_path, 'rb') as f:
                    raw_data = f.read()
                    result = chardet.detect(raw_data)
                    encoding = result['encoding']
                    text = raw_data.decode(encoding)
                return text
            except FileNotFoundError:
                Logger.error("CodeReviewCheck.test_check_code - File Not found" + file_path)
            except Exception as e:
                Logger.error("CodeReviewCheck.test_check_code - Exception : " + str(e))
                return None

        @staticmethod
        def save_to_file(content : str, file_name : str , dict_content : dict = None ):
            try :

                # 최대 열 너비 설정 (기본값: 50)
                pd.set_option('display.max_colwidth', None)

                # 생략 표시 비활성화 <- pd.set_option 표시 옵션 제어
                pd.set_option('display.max_rows', None)  # 모든 행 출력
                pd.set_option('display.max_columns', None)  # 모든 열 출력
                path = os.path.join(os.getcwd(), file_name + ".txt")
                if dict_content is not None :
                    content = json.dumps(dict_content, ensure_ascii=False, indent=4)

                with open(path, 'w') as file:
                    file.write(content)
            except Exception as e:
                Logger.error("CodeReviewCheck.CodeUI - Exception : " + str(e))

        # Code Review 결과를 DataFrame에 업데이트 : List 정보를 입력받아 없으면 OK, 있으면 추가하면서 저장
        # review_item = 코드 리뷰 항목, review_result = 코드 리뷰 결과("상세 내용", "라인", "코드")
        @classmethod
        def update_check_result(cls, review_item : str, review_result : list, df : pd.DataFrame) -> pd.DataFrame:
            try:
                # 코드 리뷰 결과에 Error 확인 -> review_result에는 NG 대상만 저장
                if len(review_result) == 0:
                    df.loc[df[COL_CR_ITEM] == review_item, COL_CR_RESULT] = ROW_CR_RESULT_OK
                else:
                    select_row = df.loc[df[COL_CR_ITEM] == review_item]
                    item_index = df.index[df[COL_CR_ITEM] == review_item].tolist()[0] + 1
                    df = df.loc[df[COL_CR_ITEM] != review_item]

                    # Error List를 Dataframe에 저장
                    for item in review_result:
                        new_row = select_row.copy()                             # ROW 정보를 새로 생성
                        new_row.loc[:, COL_CR_RESULT] = ROW_CR_RESULT_NG        # NG 저장
                        new_row.loc[:, COL_CR_LINE] = item[0]                   # 라인 위치 저장
                        new_row.loc[:, COL_CR_RESULT_DETAIL] = item[1].strip()  # 상세 내용 저장
                        new_row.loc[:, COL_CR_RESULT_CODE] = item[2].strip()    # 코드 데이터를 저장 -> Excel 파일 저장할 때만 표시

                        # 상위 부분, 새로운 행, 하위 부분 결합
                        upper_half = df.iloc[:item_index + 1]
                        lower_half = df.iloc[item_index + 1:]
                        df = pd.concat([upper_half, new_row, lower_half]).reset_index(drop=True)
                        item_index = item_index + 1

                return df

            except Exception as e:
                Logger.error("CodeReviewCheck.update_check_result - Exception : " + str(e))
        # endregion

        # lib 코드에서 코드 점검 시작
        @staticmethod
        def code_check_start( file_name : str , check_svr : bool)  :
            try:

                # 해당 데이터 테터만 삭제
                CodeReviewCheck.CodeData.init_check_list()
                #1. Load Code File
                file_path = CodeReviewCheck.CodeData.get_file_path(file_name)  # 파일 Full 경로 가져오기
                text_code = CodeReviewCheck.CodeCheck.get_file_to_text(file_path)  # 파일 코드 불러와서 문자열 변수 저장
                CodeReviewCheck.text_list = text_code.split("\n")

                # * 파일 저장
                CodeReviewCheck.CodeCheck.save_to_file(str(text_code), "[Step0] file_text")

                #2. 함수 body 정보 분리하여 저장
                CodeReviewCheck.function_body_list = CodeReviewCheck.CodeCheck.get_function_body(text_code)

                # 모두 해당되는 코드 리뷰 항목
                CodeReviewCheck.CodeCheck.code_check_loop_delay(text_code, ROW_CR_ITEM_LOOP[CR_ITEM_IDX])                   # Loop문 내에 처리 조건
                CodeReviewCheck.CodeCheck.code_check_eventminimize(text_code, ROW_CR_ITEM_EVENT_CHANGE[CR_ITEM_IDX])        # 이벤트 교환 횟수 최소화
                CodeReviewCheck.CodeCheck.code_check_callback(text_code, ROW_CR_ITEM_PROPER_DP_FCT[CR_ITEM_IDX])            # 적절한 DP 처리 함수 사용
                CodeReviewCheck.CodeCheck.code_check_UnnecessaryCode(text_code, ROW_CR_ITEM_UNNECESSARY_CODE[CR_ITEM_IDX])  #
                CodeReviewCheck.CodeCheck.code_check_hard_coding(text_code, ROW_CR_ITEM_HARD_CODE[CR_ITEM_IDX])




                # Code + 코드 리뷰 아이템 정보 전달 -> 코드 리뷰  진행 후 해당 DataFrame에 결과 저장하여 반환
                if (check_svr == True):
                    Logger.info("CodeCheck.code_check_start(SVR)")

                    # SVR에서만 점검하는 코드 리뷰 항목
                    CodeReviewCheck.CodeCheck.code_check_dp_exception(text_code, ROW_CR_ITEM_DP_EXCEPTION[CR_ITEM_IDX])     # DP 함수 예외 처리
                    CodeReviewCheck.CodeCheck.code_check_try_exception(text_code, ROW_CR_ITEM_TRY_EXCEPTION[CR_ITEM_IDX])   # try, Catch 예외 처리
                    CodeReviewCheck.CodeCheck.code_check_version(text_code, ROW_CR_ITEM_VERSION[CR_ITEM_IDX])               # 버전 정보 작성 확인인


                else:
                    Logger.info("CodeCheck.code_check_start(CLI)")
                    # Client에서만 점검하는 코드 리뷰 항목

                # return CodeReviewCheck.df_crc_result
            except Exception as e:
                Logger.error("CodeReviewCheck.test_check_code - Exception : " + str(e))

        # [코드 표준] 불필요한 코드 지양 -> SVR
        @classmethod
        def code_check_UnnecessaryCode(cls, text_code : str, cr_item : str) :
            def remove_line_comments(code:str) -> str:
                try:
                    lines = code.splitlines()
                    modified_lines = []
                    for line in lines:
                        # Remove comments starting with //
                        if '//' in line:
                            line = line.split('//')[0] + ' ' * len(line.split('//')[1])
                        if '#' in line:
                            line = line.split('#')[0] + ' ' * len(line.split('#')[1])
                        modified_lines.append(line)

                    return '\n'.join(modified_lines)
                except Exception as e:
                    Logger.error("CodeCheck.remove_line_comments - Exception : " + str(e))

            def find_variable_usage(code: str, variable: str) -> bool:
                try :
                    # Use a word boundary to ensure exact match
                    pattern = re.compile(r'\b' + re.escape(variable) + r'\b')
                    matches = pattern.findall(code)
                    return len(matches) > 1  # True if more than one occurrence (considering the declaration line)
                except Exception as e:
                    Logger.error("CodeReviewCheck.test_check_code - Exception : " + str(e))

            try:
                Logger.info("CodeCheck.code_check_UNUSED - Start")

                # 1. 코드에서 주석 부분 제거 (공백으로 변경, 라인 수 유지 필요)
                # new_text_code = remove_line_comments(text_code)

                # 2. 전역 변수 찾기
                global_vars = cls.get_variables(text_code)

                # 3. 전역 변수 사용 되었는지 확인
                total_error_result = []
                for var_name, line in global_vars :
                    used_flag = False
                    body_code = ""
                    for item in CodeReviewCheck.function_body_list:
                        function_name = item[0]
                        body_code = item[1]
                        # body_code = remove_line_comments(body_code)
                        start_number = item[2]

                        # local_vars = cls.get_variables(body_code)
                        #
                        # # #3.1 지역 변수 사용 확인
                        # for local_var_name, local_line in local_vars :
                        #     if find_variable_usage(body_code, local_var_name ) == False :
                        #         total_error_result = total_error_result + [[start_number + local_line, f"{function_name}함수의 {local_var_name} 변수가 사용 이력이 없습니다."]]

                        #3.2 전역 변수 사용 확인
                        if find_variable_usage(body_code, var_name) == True :
                            Logger.debug(f"CodeCheck.code_check_UnnecessaryCode - Find OK. funtion = {function_name, start_number}, used_var = {var_name}" )
                            used_flag = True
                            continue

                    if not used_flag:
                        total_error_result = total_error_result + [[line, f"{var_name} 변수가 사용 이력이 없습니다.", ""]]

                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, total_error_result, CodeReviewCheck.df_crc_result)

            except Exception as e:
                Logger.error("CodeCheck.test_check_code - Exception : " + str(e))

        # 전역 변수 리스트에 저장
        @classmethod
        def get_variables(cls, text_code : str) -> list:
            try :
                lines = text_code.splitlines()
                global_vars = []
                brace_depth = 0

                for i, line in enumerate(lines):
                    stripped_line = line.strip()

                    # Update brace depth
                    brace_depth += stripped_line.count('{')
                    brace_depth -= stripped_line.count('}')

                    if brace_depth == 0:  # We're not inside any function or block
                        # Split the line at '=' and consider only the left side
                        left_side = re.split(r'=', stripped_line, 1)[0]

                        # Check for exclusion condition: single word without ',', '=', ';'
                        if ',' not in left_side and '=' not in stripped_line and ';' not in stripped_line:
                            continue

                        # Further split by delimiters ',', ';' and process each part
                        parts = re.split(r'[;,]', left_side)
                        for part in parts:
                            part = part.strip()
                            # Ignore parts containing parentheses
                            if '(' in part or ')' in part:
                                continue
                            if part:
                                # Split by whitespace and take the last element as the variable name
                                words = part.split()
                                var_name = words[-1] if words else ''

                                # Remove any trailing special characters and check if it's a valid identifier
                                var_name = re.sub(r'[^a-zA-Z0-9_]', '', var_name)

                                # Exclude numeric-only variables and ensure it starts with an alphabetic character
                                if var_name and not var_name.isdigit() and re.match(r'^[a-zA-Z_]\w*$', var_name):
                                    global_vars.append((var_name, i + 1))  # Store name and line number

                return global_vars
            except Exception as e:
                Logger.error("CodeReviewCheck.get_variables - Exception : " + str(e))

        # [코드 표준] 스크립트 이력 관리 -> SVR
        @classmethod
        def code_check_version(cls, text_code : str, cr_item : str)  :
            try :
                #1. 버전 이력 관리 변수 확인
                version_variable_list = ["g_script_release_version", "g_script_release_date"]
                check_result = True
                detail_msg = ""
                for list_item in version_variable_list :
                    return_code = cls.is_variable_used(text_code, list_item)
                    # return_code = CodeReviewCheck.CodeCheck.is_variable_used(text_code, list_item)

                    if len(return_code) == 0 :
                        Logger.error("CodeReviewCheck.code_check_version - Check NG. " + list_item)
                        detail_msg = detail_msg + f"{list_item} 선언 되지 않았습니다. \n"
                        check_result = False
                    else :
                        detail_msg = detail_msg + f"{list_item} 선언 되었습니다. \n"
                        Logger.info("CodeReviewCheck.code_check_version - Check OK. " + list_item)

                # DataTable 업데이트
                if check_result == True :
                    CodeReviewCheck.df_crc_result.loc[CodeReviewCheck.df_crc_result[COL_CR_ITEM] == cr_item, COL_CR_RESULT] = ROW_CR_RESULT_OK
                if check_result == False :
                    CodeReviewCheck.df_crc_result.loc[CodeReviewCheck.df_crc_result[COL_CR_ITEM] == cr_item, COL_CR_RESULT] = ROW_CR_RESULT_NG
                CodeReviewCheck.df_crc_result.loc[CodeReviewCheck.df_crc_result[COL_CR_ITEM] == cr_item, COL_CR_LINE] = "-"
                CodeReviewCheck.df_crc_result.loc[CodeReviewCheck.df_crc_result[COL_CR_ITEM] == cr_item, COL_CR_RESULT_DETAIL] = detail_msg
                CodeReviewCheck.df_crc_result.loc[CodeReviewCheck.df_crc_result[COL_CR_ITEM] == cr_item, COL_CR_RESULT_CODE] = "-"

            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_version - Exception : " + str(e))

        # [코드 표준] 하드 코딩 지양 -> 공통
        @classmethod
        def code_check_hard_coding(cls, text_code : str, cr_item : str):
            try :
                Logger.info("CodeCheck - code_check_hard_coding start")

                # 1. body code 에서 HardCoding 부분 검출 및 결과 df_crc_result 업데이트
                total_error_result = []
                for item in CodeReviewCheck.function_body_list :
                    hard_coding_list = cls.get_hard_coding_check(item)
                    total_error_result = total_error_result + hard_coding_list

                # 2. 코드 리뷰 결과 Dataframe에 업데이트 -> CodeReviewCheck.df_crc_result
                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, total_error_result, CodeReviewCheck.df_crc_result)
            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        #[코드 표준] function의 Body를 스캔하여 Hard Coding 부분을 리스트에 저장
        @classmethod
        def get_hard_coding_check(cls, function_info : list) -> list:
            try:
                function_name = function_info[0]
                body_code = function_info[1]
                start_number = function_info[2]
                result_hard_coding = []   # Line 위치, 상세 내용

                # 하드 코딩 패턴 1 : 대입 연산자 하드 코딩 체크
                line_count = start_number
                for line_code in body_code.split('\n') :
                    # Skip 대상이 아니고 대입 연산자 경우 하드 코딩이 있는 경우
                    if cls.check_skip_string(line_code) == False and cls.is_check_hard_coding_opeartion(line_code) == True:
                        # result_hard_coding.append([line_count+2, "하드 코딩으로 작성되었습니다. 함수 = %s, 코드 = %s" % (function_name, line_code.strip() )])
                        result_hard_coding.append([line_count + 2, f"하드 코딩으로 작성되었습니다. 함수 = {function_name}", line_code.strip()])
                    line_count = line_count + 1

                # 하드 코딩 패턴 2 : 함수 or 괄호 안에 하드 코딩이 설정 되어 있는 경우
                pattern2_hard_coding_list = cls.is_check_hard_coding_blacket(function_name, body_code, start_number)
                result_hard_coding = result_hard_coding + pattern2_hard_coding_list

                return result_hard_coding
            except Exception as e :
                Logger.error("CodeReviewCheck.get_hard_coding - Exception : " + str(e))

        @classmethod
        def is_check_hard_coding_opeartion(cls, line_code:str) -> bool:
            try:
                r_pattern = re.compile(r'=\s*([^;]+)\s*;')  # [Pattern] 대입 연산자에서 우항 캡쳐
                non_operaion_pattern = re.compile(r'[^\s\+\-\*/\%]+')  # [Pattern] 대입 연산자에서 우항 캡쳐
                match = r_pattern.search(line_code)

                result = False
                if match:
                    # 1. 우항 데이터 캡쳐
                    r_expression = match.group(1).strip()
                    # 2. 우항 데이터에서 비연산자들만 리스트로 반환
                    non_operators = non_operaion_pattern.findall(r_expression)
                    # 3. 문자열 or 숫자만 있는 경우 체크
                    for item in non_operators:
                        if cls.is_hard_coding_check(item) == True:
                            result = True
                            break

                return result

            except Exception as e:
                Logger.error("CodeReviewCheck.is_check_hard_coding_opeartion - Exception : " + str(e))

        @classmethod
        def is_hard_coding_check(cls, line_code:str) -> bool:
            if line_code.isdigit() or '"' in line_code:
                return True
            else :
                return False

        @classmethod
        def is_check_hard_coding_blacket(cls, function_name:str,  body_code:str, body_start_number : int) -> list:
            try:
                # 1. 괄호안의 내용 추출
                # 정규식 패턴: 괄호 안의 내용을 추출 (멀티라인 지원, 비연산자 포함)
                pattern = re.compile(r'\(([^)]*)\)', re.DOTALL)
                matches = pattern.finditer(body_code)

                # 2. body Code를 한 줄씩 저장
                line_code = body_code.split('\n')

                # 3. 하드 코딩 부분 리스트에 저장하여 반환
                result = []
                for match in matches:
                    # 공백이나 빈 인자도 포함하여 캡쳐
                    parts = re.split(r'\s*[\+\-\*/\%,]\s*', match.group(1))                         # 패턴에 따라서 분리
                    non_operators = [part.strip() for part in parts if part.strip() or part == ""]  # 분리된 패턴을 공백을 제거하여 리스트로 저장
                    start_line = body_code.count('\n', 0, match.start()) + 1    # 매칭된 그룹의 시작 줄 번호를 계산
                    original_line = line_code[start_line - 1].strip()           # 원본 문자열을 포함하여 저장

                    if (cls.check_skip_string(original_line) == False): # skip 대상 있는지 확인
                        for operator_item in non_operators:
                            if cls.is_hard_coding_check(operator_item) == True: # 비연산자를 하나씩 하드코딩(문자열 or 상수) 되어 있는지 확인
                                result.append((body_start_number + start_line + 1, "하드 코딩으로 작성되었습니다. 함수 = %s, 코드 = %s" % (function_name, original_line.strip())))
                                break
                return result
            except Exception as e:
                Logger.error("CodeReviewCheck.is_check_hard_coding_blacket - Exception : " + str(e))

        @classmethod
        def check_skip_string(cls, input_text: str) -> bool:
            try:
                # Skip 대상이 있는 경우 True로 반환
                skip_list = ['Debug', 'dpConnect', 'writeLog', 'startThread', 'update_user_alarm', 'read_config', 'paCfg', 'for', 'sprintf', 'FROM', 'WHERE']
                skip_check = False
                skip_check = any(list_item in input_text for list_item in skip_list)

                return skip_check

            except Exception as e:
                Logger.error("CodeReviewCheck.check_skip_string - Exception : " + str(e))


        # [코드 표준] Try, Catch 예외 처리 -> 공통
        @classmethod
        def code_check_try_exception(cls, text_code, cr_item:str):
            try:
                # 1 함수, body 정보 추출
                # function_list = cls.extract_functions_from_code(text_code)
                # save_text = ""
                # for item in function_list :
                #     save_text = save_text + str(item[0])  + "\n"
                #
                # CodeReviewCheck.CodeCheck.save_to_file(str(save_text), "[Step0] code_check_try_exception")

                # 2. 함수의 body 코드를 전달하여 try, catch 설정 확인
                # 설정 안된 함수의 경우 list에 저장 -> result_try_exception
                result_try_exception = []
                for item in CodeReviewCheck.function_body_list :
                    function_name = item[0]
                    body_code = item[1]
                    start_number = item[2]

                    if cls.is_try_exception(body_code) == False :
                        result_try_exception.append([start_number, f"함수에 try,Catch 예외 처리가 누락 되었습니다.  함수 = {function_name}", ""])

                # 3. 코드 리뷰 결과 Dataframe에 업데이트 -> CodeReviewCheck.df_crc_result
                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, result_try_exception, CodeReviewCheck.df_crc_result)

            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        @classmethod
        def is_try_exception(cls, body_code : str) -> bool :
            try :
                # 정규 표현식 패턴 정의
                pattern = r'\s*try\s*\{.*?\}\s*(//.*\s*)?catch\s*\{.*?\}'
                # 정규 표현식 컴파일
                regex = re.compile(pattern, re.DOTALL)
                #패턴이 텍스트에 있는지 확인
                match = regex.search(body_code)
                if match is not None :
                    return True
                else :
                    return False
            except Exception as e:
                Logger.error("CodeReviewCheck.is_try_exception - Exception : " + str(e))


        # [코드 표준] DP 함수 예외 처리 -> 공통
        @classmethod
        def code_check_dp_exception(cls, text_code : str,  cr_item : str):
            try:
                # DP 함수 예외 처리 패턴 : dp*로 시작 하는 함수에서 대입 연산자가 없는 패턴 찾기
                pattern = r'(?<![!=\s])\s*dp[a-zA-Z][a-zA-Z0-9_]*\([^)]*\)\s*(?![!=\s])'
                total_error_result = []

                for  item in CodeReviewCheck.function_body_list :
                    fnc_name = item[0]
                    body_code = item[1]
                    start_line = item[2]
                    # end_line = item[3]
                    error_list = cls.get_pattern(body_code, pattern, start_line)
                    for error_item in error_list :
                        detail_msg = f"DP 함수 예외 처리가 되지 않았습니다. 함수 = {fnc_name}, Code = {error_item[1]}"
                        total_error_result = total_error_result + [[error_item[0], detail_msg]]

                # 3. 코드 리뷰 결과 Dataframe에 업데이트 -> CodeReviewCheck.df_crc_result
                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, total_error_result, CodeReviewCheck.df_crc_result)
            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        @classmethod
        def get_pattern(cls, text:str, pattern:str, function_line : int = 0) -> list:

            result_list = []

            try:
                matches = re.finditer(pattern, text)

                for match in matches :
                    start_pos = match.start()
                    match_text = match.group(0)
                    match_line_count = match_text.count('\n')
                    line_number = text.count('\n',0, start_pos) + function_line + match_line_count
                    result_list = result_list + [[line_number, match_text.strip()]]

            except Exception as e:
                Logger.error("CodeReviewCheck.get_pattern - Exception : " + str(e))

            return result_list


        # [성능] 스크립트 동작 Active 감시 적용
        @classmethod
        def code_check_active(cls, text_code):
            try:
                None

            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        # [성능] Loop문 내 처리조건 확인
        @ classmethod
        def code_check_loop_delay(cls, text_code : str, cr_item : str):
            try:
                Logger.info("CodeCheck - code_check_loop_delay start")
                total_error_result = []
                for index, item in enumerate(CodeReviewCheck.function_body_list) :
                    fnc_name = item[0]
                    body_code = item[1]
                    start_line = item[2]
                    end_line = item[3]
                    while_delay_miss = True
                    
                    #while문자열을 제외하고 내부에 있는 코드를 반환
                    while_code = cls.get_while_code(body_code)      # while문 코드를 확인하여 반환
                    detail_msg = ""

                    if len(while_code) > 0 and cls.is_check_while_delay(while_code) == False :
                        detail_msg = f"while문 내에서 delay 코드가 작성되지 않았습니다. 함수 = {fnc_name}"
                        total_error_result = total_error_result + [[start_line, detail_msg, ""]]

                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, total_error_result, CodeReviewCheck.df_crc_result)
            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        # [성능] Loop문 내 처리조건 확인 -> Function의 Body 코드에서 while문 코드만 저장
        @classmethod
        def get_while_code(cls, body_code: str) -> str:

            try :
                pattern = r'while\s*\([^)]*\)\s*{'
                match = re.search(pattern, body_code, re.DOTALL)
                while_block_code = ""

                if match:
                    while_code_start_pos = match.end() - 1
                    end_index = cls.find_closing_brace_index(body_code, while_code_start_pos)
                    if end_index == -1:
                        while_block_code = ""
                    else:
                        while_block_code = body_code[while_code_start_pos + 1: end_index]

                else:
                    while_block_code = ""

                return while_block_code
            except Exception as e:
                Logger.error("CodeReviewCheck.get_while_code - Exception : " + str(e))

        # [성능] Loop문 내 처리조건 확인 -> Function의 Body 코드에서 while문 코드만 저장
        # Flase -> 누락, True -> 확인
        @classmethod
        def is_check_while_delay(cls, body_code: str) -> bool:
            try :
                check_keyward = "delay"
                check_result = False
                pattern = re.compile(r'([^\s].*?)\s*(?=\{)')  # "if() {" 을 찾는 패턴
                matches = list(pattern.finditer(body_code))

                # while문 내에서 {}으로 설정되어 있는지 확인하여 내부 코드는 삭제
                for match in matches:
                    block_name = match.group(1)

                    # block 조건문으로 판단하는 경우 코드는 제외
                    if block_name not in "finally":
                        block_start_pos = match.end()  # '{' 포함 시켜서 index 전달
                        block_end_pos = cls.find_closing_brace_index(body_code, block_start_pos)  # 블록 마지막 위치를 반환
                        body_code = cls.replace_with_spaces(body_code, block_start_pos + 1, block_end_pos)  # 블록 시작 위치와 마지막 위치 코드는 삭제 (시작위치는 '{' 제외하고 전달)

                if check_keyward in body_code:
                    check_result = True
                return check_result

            except Exception as e:
                Logger.error("CodeReviewCheck.is_check_while_delay - Exception : " + str(e))

        # [성능] 이벤트 교환 횟수 최소화
        @ classmethod
        def code_check_eventminimize(cls, text_code:str, cr_item:str):
            try:
                Logger.info("CodeCheck - code_check_eventminimize start")
                total_error_result = []
                for index, item in enumerate(CodeReviewCheck.function_body_list) :
                    fnc_name = item[0]
                    body_code = item[1]
                    start_line = item[2]
                    end_line = item[3]
                    loop_error_list = cls.loop_pattern_check(body_code, start_line)
                    for error_item in loop_error_list :
                        detail_msg = f"DP 함수가 반복적으로 처리 되었습니다. 함수 = {fnc_name} "
                        total_error_result = total_error_result + [[error_item[1], detail_msg, error_item[0]]]

                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, total_error_result, CodeReviewCheck.df_crc_result)
            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        @classmethod
        def loop_pattern_check(cls, body_code:str, start_line:int) -> list:
            try:
                # 정규식 패턴을 입력 받아 입력 텍스트에 패턴이 있는지 확인하는 함수
                def check_pattern(loop_block_code: str, pattern_text: str, ) -> bool:
                    # 1. 정규식 패턴: dp로 시작하고 대문자 문자로 이어진 후 (로 끝나는 패턴
                    # pattern = re.compile(r'(dp[A-Z]\w*\()', re.MULTILINE)
                    pattern = re.compile(pattern_text, re.MULTILINE)
                    matches = pattern.finditer(loop_block_code)
                    found = False

                    # 2. dp* 함수가 있는 경우 체크
                    for match in matches:
                        return  True
                    else :
                        return False

                # 1. for문의 정규식 생성
                pattern = re.compile(r'(for\s*\(.*?\)\s*\{[^}]*\})', re.DOTALL)

                # 2. 정규식에 맞는 코드 매칭
                matches = pattern.finditer(body_code)

                # 3. 리턴할 데이터 리스트 저장(에러 반복문) : string, int -> result_list
                result_list = []
                lines = body_code.splitlines()

                # 4. for문 패턴 매칭 동작
                for match in matches:
                    match_text = match.group(0)  # 캡쳐 되는 첫번째 그룹
                    match_start = match.start()

                    # Calculate line number
                    line_number = body_code[:match_start].count('\n') + 1

                    # for문 Block에 dp*패턴이 있는지 확인
                    dp_function_pattern = r'(dp[A-Z]\w*\()'
                    if check_pattern(match_text, dp_function_pattern) == True:

                        # title 찾는 정규식
                        title_match = re.match(r'(for\s*\(.*?\))', match_text)

                        # loop문 title와 라인 number 리스트로 저장 -> result_list
                        if title_match:
                            title = title_match.group(0)
                            result_list = result_list + [[title, start_line + line_number]]
                        else:
                            result_list = result_list + [[title_match, start_line + line_number]]
                return result_list

            except Exception as e:
                Logger.error("CodeReviewCheck.loop_pattern_check - Exception : " + str(e))

        # [성능] 적절한 DP 함수 사용
        @ classmethod
        def code_check_callback(cls, text_code:str, cr_item:str):
            try:
                connection_fnc_list = cls.get_dpconnect_list(text_code)
                total_error_result = []
                for function_item in CodeReviewCheck.function_body_list:
                    fnc_name = function_item[0]
                    body_code = function_item[1]
                    start_line = function_item[2]
                    end_line = function_item[3]

                    for conn_function_item in connection_fnc_list :
                        if fnc_name == conn_function_item :
                            delay_pattern = r'delay\s*\(.*?\)\s*;'
                            matches = re.findall(delay_pattern, body_code)
                            # delay가 있는 경우에는 에러 내용 저장
                            if bool(matches) == True :
                                detail_msg = f"Callback 함수에서 delay 코드가 존재합니다. 함수 = {fnc_name} "
                                total_error_result = total_error_result + [[start_line, detail_msg, str(matches)]]
                CodeReviewCheck.df_crc_result = cls.update_check_result(cr_item, total_error_result, CodeReviewCheck.df_crc_result)
            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_callback - Exception : " + str(e))

        # dpQueryConnect or dpConnect 사용하는 Connection 함수를 리스트에 저장 (중복 제거)
        @classmethod
        def get_dpconnect_list(cls, body_text: str) -> list:
            try :
                # connect Call 하는 코드에서 Connection 함수 이름을 리턴
                def get_function_name(conncet_code_text: str) -> str:
                    # dpConnect("", ...) 에서 첫번째 "" 문자열을 캡쳐
                    match = re.search(r'"(.*?)"', conncet_code_text)
                    if match:
                        return match.group(1)
                    else:
                        return ""

                result_list = []  # 함수에서 connect 함수만 저장하여 반환

                # dp*Connect*(*); 패턴으로 코드 검색
                pattern = r'dp\w*Connect\w*\s*\(\s*[\s\S]*?\s*\)\s*;'  # dp*Connect* 함수 패턴
                matches = re.findall(pattern, body_text, re.DOTALL)  # re.DOTALL 옵션은

                # dp*Connect*(*) 코드를 '\n' 제외하고 리스트에 저장
                connect_pattern_list = [re.sub(r'\s+', ' ', match) for match in matches]        # 줄바꿈 문자 제거 후 결과 출력

                # 코드에서 Callback하는 함수를 찾아서(get_function_name) 리스트에 저장
                for item in connect_pattern_list:
                    connect_function_name = get_function_name(item)
                    if len(connect_function_name) > 0:
                        result_list = result_list + [connect_function_name]

                return result_list
            except Exception as e:
                Logger.error("CodeReviewCheck.get_dpconnect_function - Exception : " + str(e))

        # [성능] Raima DB 증가 방지
        @ classmethod
        def code_check_raimaDB(cls, text_code):
            try:
                None

            except Exception as e:
                Logger.error("CodeReviewCheck.code_check_hardcoding - Exception : " + str(e))

        @classmethod
        def extract_functions_from_code2(cls, text:str) -> list:
            try:
                # function_dict = {}

                # 함수 이름과 본문을 찾는 정규 표현식
                # 주석 제외, 타입, 함수, 괄호 매개변수, { } body 구분
                pattern = r'(?<!/)\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\([^)]*\)\s*\{\n(.*?)\n\}'
                # pattern = r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*\([^)]*\)\s*\{\n(.*?)\n\}'
                # pattern = r'\b([a-zA-Z_][a-zA-Z0-9_]*)\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*\([^)]*\)\s*\{\n(.*?)\n\}'
                # pattern = r'\b(?:[a-zA-Z_][a-zA-Z0-9_]*)\s+([a-zA-Z_][a-zA-Z0-9_]*)\s*\([^)]*\)\s*\{\n(.*?)\n\}'

                # matches에서는 1번 캡처는 함수이름 2번 캠처는 body 부분입니다.
                matches = re.finditer(pattern, text, re.DOTALL)
                results = []

                # (1) 함수 이름, (2) BodyCode, (3) 함수 시작 라인 수, (4) 함수 종료 라인 수 정보를 리스트에 저장
                for match in matches:
                    function_name = match.group(1)
                    function_body = match.group(2)
                    start_pos = match.start()
                    end_pos = match.end()
                    start_line = cls.calculate_line_number(text, start_pos)
                    end_line = cls.calculate_line_number(text, end_pos)
                    results.append([function_name, function_body, start_line, end_line])

                # return function_dict
                return results
            except Exception as e:
                Logger.error("CodeReviewCheck.extract_functions_from_code - Exception : " + str(e))

        # 함수 body 부분 추출 : body code, body start line, body end line
        @classmethod
        def get_body_info(cls, text : str) -> list:
            try:
                stack = []
                result = []
                line_number = 1
                start_index = None
                start_line = None

                for i, char in enumerate(text):
                    if char == '\n':
                        line_number += 1
                    elif char == '{':
                        if start_index is None:
                            start_index = i
                            start_line = line_number
                        stack.append((i, line_number))
                    elif char == '}' and stack:
                        start_pos, start_ln = stack.pop()
                        if not stack:
                            end_line = line_number
                            result = result + [[text[start_index + 1:i], start_ln, end_line]]
                            start_index = None
                            start_line = None

                return result
            except Exception as e:
                Logger.error("CodeReviewCheck.get_body_info - Exception : " + str(e))

        @classmethod
        def get_function_name(cls, text: str) -> str:
            try :
                # pattern = r'(?<!/)\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\('
                pattern = r'(?<!/)\b([a-zA-Z_][a-zA-Z0-9_]*)\s*\('

                matches = re.finditer(pattern, text)
                func_name = ""
                for match in matches:
                    func_name = match.group(1)
                    match_start_pos = match.start()

                    # 주석의 경우는 제외
                    l_pos = text.find('/')
                    if l_pos >= 0 and l_pos < match_start_pos:
                        func_name = ""
                    else :
                        break

                return str(func_name)
            except Exception as e:
                Logger.error("CodeReviewCheck.get_function_name - Exception : " + str(e))

        @classmethod
        def get_function_body(cls, text):
            try:
                Logger.info("CodeCheck - get_function_body start")
                function_info = []

                # 주석 코드는 공백으로 삭제
                text = cls.remove_line_comments(text)

                # 1. 먼저 body의 정보를 먼저 저장 : body_code, body_start_line, body_end_line
                body_info = cls.get_body_info(text)
                # for index, item in enumerate(body_info):
                #     print(index, item[1], item[2])

                # 2. function과 body를 분리하여 리스트에 저장 -> 2차원 배열 function 이름, body, start_line, end_line
                end_line = 0
                # for index, code_text in enumerate(code_texts) :
                code_line_number = 0
                for code_text in text.split('\n'):
                    code_line_number = code_line_number + 1

                    # 3. 정규식으로 함수이름 조회
                    if code_line_number >= end_line:
                        func_name = cls.get_function_name(code_text)

                    if 0 < len(func_name):
                        body_item = body_info.pop(0)
                        body_code = body_item[0]
                        start_line = body_item[1]
                        end_line = body_item[2]

                        # 함수 이름, 함수 body, 함수 body 시작 라인, 함수 body 마지막 라인
                        function_info = function_info + [[func_name, body_code, start_line, end_line]]
                        func_name = ""

                return function_info
            except Exception as e:
                Logger.error("CodeReviewCheck.extract_functions_from_code - Exception : " + str(e))

        @classmethod
        def calculate_line_number(cls, text, index):
            try :
                return text.count('\n', 0, index) + 1
            except Exception as e:
                Logger.error("CodeReviewCheck.calculate_line_number - Exception : " + str(e))

        @classmethod
        def remove_comments(cls, code: str) -> str:
            try:
                # 라인 주석 제거
                code = re.sub(r'//.*', '', code)
                # 블록 주석 제거
                code = re.sub(r'/\*.*?\*/', '', code, flags=re.DOTALL)

                return code
            except Exception as e:
                Logger.error("CodeReviewCheck.remove_comments - Exception : " + str(e))

        # 전역 변수 찾기
        @staticmethod
        def extract_global_variables(code: str) -> list:
            try:

                # . 중괄호 내용(함수 정의 등)을 제외하고 전역 영역의 코드만 추출
                # 1. 주석 삭제
                code = __class__.remove_comments(code)

                # 2. 함수 영여 제거
                code_without_braces = __class__.extract_global_scope_code(code)
                # code_without_braces = re.sub(r'{[^{}]*}', '', code_without_braces)

                # 3. 라인 별로 처리하여 주석 제거
                lines = code_without_braces.split('\n')
                lines_without_comments = [re.sub(r'//.*', '', line).strip() for line in lines if
                                          '//' in line or line.strip() != '']


                # 전역 변수를 찾기 위한 정규식 패턴: 세미콜론으로 끝나는 모든 선언 찾기
                pattern = re.compile(r'\b(\w+)\s*([^;]+);')

                global_variables = []

                for line in lines_without_comments:
                    matches = pattern.finditer(line)
                    for match in matches:
                        # 변수 이름만 추출
                        variables_part = match.group(2)
                        # 쉼표로 구분된 여러 변수 처리
                        variables = [var.split('=')[0].strip() for var in variables_part.split(',')]
                        global_variables.extend(variables)

                return global_variables
            except Exception as e:
                Logger.error("CodeReviewCheck.extract_global_variables - Exception : " + str(e))

        # 전역 코드 영역만 찾기
        @classmethod
        def extract_global_scope_code(cls, code: str) -> str:
            # 괄호 안의 내용(지역 영역)을 제외하고 전역 영역의 코드만을 추출
            nested_braces_pattern = re.compile(r'{[^{}]*}')
            while re.search(nested_braces_pattern, code):
                code = re.sub(nested_braces_pattern, '', code)
            return code

        # 코드에서 변수 선언 확인
        @classmethod
        def is_variable_used(cls, script_code : str, variable_name : str) -> str:
            try:
                # \b : 단어 경계, escape : 특수문자 이스케이프 처리, \s* : 변수 이름 뒤 공백, (?:=|;|,) : "=",  ";", "," 문자 포함
                pattern = re.compile(r'\b' + re.escape(variable_name) + r'\b\s*(?:=|;|,)')

                matches = re.search(pattern, script_code)
                if matches :
                    return matches.group(0)
                else :
                    return ""
                # Search for the pattern in the code
                # if re.search(pattern, script_code):
                #     return True
                # return False

            except Exception as e:
                Logger.error("CodeReviewCheck.extract_global_variables - Exception : " + str(e))

        # 변수의 값 확인
        def get_variable_value(c_code, variable_name):

            #1. 전달한 변수 정규식 생성 :
            pattern = re.compile(r'\b' + re.escape(variable_name) + r'\b\s*=\s*("[^"]*"|\d+|[^,;]+)')

            # Search for the pattern in the code
            match = re.search(pattern, c_code)
            if match:
                return match.group(1).strip()
            return None

    @staticmethod
    def add_crc_info(new_dict):
        try:
            pass
            # 1. UI에서 선택한 파일


        except Exception as e:
            Logger.error("CodeReviewCheck.add_crc_info -  Exception : " + str(e))
