import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows

# 예제 데이터 생성
data = {
    'Group': ['1', '2', '2'],
    'Name': ['John', 'Alice', 'Bob'],
    'Age': [30, 25, 35],
    'City': ['New York', 'Los Angeles', 'Chicago']

}

# DataFrame 생성
df = pd.DataFrame(data)

# 엑셀 워크북 생성
wb = Workbook()
ws = wb.active




for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# 모든 셀 가운데 정렬 및 맞춤
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 병합할 셀 영역
merge_range = 'A3:A4'

# 병합된 셀의 값을 첫 번째 셀의 값으로 설정
first_cell = ws['A3']
value_to_merge = first_cell.value
ws.merge_cells(merge_range)

# 병합된 셀에 값 설정
merged_cell = ws['A3']
merged_cell.value = value_to_merge

# 병합된 셀 가운데 정렬 및 맞춤
merged_cell.alignment = Alignment(horizontal='center', vertical='center')


# 가운데 맞춤
# 셀 병합


# 엑셀 파일 저장
current_time = datetime.now()
format_time = current_time.strftime("%Y%m%d%H%M%S")
file_name = "output_%s.xlsx" % (format_time)
# print(file_name)
wb.save(file_name)