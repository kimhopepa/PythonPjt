from openpyxl import Workbook
from openpyxl.styles import Font

# 워크북과 시트 생성
wb = Workbook()
ws = wb.active

# 열 비율과 총 너비를 기반으로 열 너비 조정
def adjust_column_widths(ws, column_ratios, total_width):
    # 열 비율 검증
    if len(column_ratios) != len(list(ws.columns)):
        raise ValueError("열 비율의 개수와 열의 개수가 일치 해야 합니다.")

    # # 각 열의 문자 정의
    columns = [cell[0].column_letter for cell in ws.columns]

    # # 열 비율에 따른 너비 조정
    for column, ratio in zip(columns, column_ratios):
        adjusted_width = total_width * ratio
        ws.column_dimensions[column].width = adjusted_width

def change_text_color(ws, target_text, target_column, color="FF0000"):
    """
    ws: 워크시트
    target_text: 변경할 텍스트
    target_column: 텍스트 색상을 변경할 컬럼 (예: 'B')
    color: 색상 코드 (기본값: 빨간색)
    """
    for row in ws.iter_rows():
        cell = row[ws[target_column + '1'].column - 1]  # 해당 열의 셀 가져오기
        if target_text in str(cell.value):
            cell.font = Font(color=color)

if __name__ == '__main__':
    print("start")

    data = [
        ["Name", "Address", "Phone Number", "Email", "DOB", "Occupation", "City"],
        ["Alice", "123 Apple St", "555-1234", "alice@example.com", "1990-01-01", "Engineer", "New York"],
        ["Bob", "456 Banana Ave", "555-5678", "bob@example.com", "1985-05-12", "Doctor", "Los Angeles"],
        ["Charlie", "789 Cherry Blvd", "555-8765", "charlie@example.com", "1992-09-23", "Artist", "Chicago"],
        ["David", "1010 Pine St", "555-1010", "david@example.com", "1988-12-30", "Lawyer", "San Francisco"],
    ]
    # 예제 데이터 입력
    # data = [
    #     ["Name", "Address", "Phone Number"],
    #     ["Alice", "123 Apple St", "555-1234"],
    #     ["Bob", "456 Banana Ave", "555-5678"],
    #     ["Charlie", "789 Cherry Blvd", "555-8765"],
    # ]

    for row in data:
        ws.append(row)
    # # 예를 들어, 100을 총 너비로 설정 (적절한 값으로 조정 필요)
    # # 각 열에 대한 비율 (합계가 1이어야 함)
    column_ratios = [0.4, 0.3, 0.3, 0.2, 0.2, 0.2, 0.2]  # 각 열의 비율 설정 (Name, Address, Phone Number)
    adjust_column_widths(ws, column_ratios, 100)
    change_text_color(ws, "Engineer", target_column="F")
    #
    # # 워크북 저장
    wb.save("adjusted_width_with_ratios.xlsx")