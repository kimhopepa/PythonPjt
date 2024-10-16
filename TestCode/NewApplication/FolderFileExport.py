import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def normalize_path(path):
    # 원시 문자열로 변환
    normalized_path = os.path.normpath(path)
    # 마지막에 백슬래시가 있으면 제거
    if normalized_path.endswith(os.sep):
        normalized_path = normalized_path[:-1]
    return normalized_path

def list_files_and_folders(path):
    # 경로 정규화
    path = normalize_path(path)

    # 경로가 유효한지 확인
    if not os.path.exists(path):
        raise ValueError("입력된 경로가 존재하지 않습니다.")

    # 결과를 저장할 리스트 초기화
    file_folder_list = []

    # 경로와 하위 폴더의 파일 및 폴더 목록 가져오기
    for root, dirs, files in os.walk(path):
        # 현재 폴더 이름 추가
        folder_name = os.path.basename(root)
        file_folder_list.append({"폴더 이름": folder_name, "파일 이름": "", "파일 전체 경로": "", "파일 개수": len(files)})

        # 파일 이름 추가
        for file in files:
            file_path = os.path.join(root, file)
            file_folder_list.append({"폴더 이름": folder_name, "파일 이름": file, "파일 전체 경로": file_path, "파일 개수": 0})  # 파일 개수는 0으로 설정

    # DataFrame으로 변환
    df = pd.DataFrame(file_folder_list)

    return df

def save_to_excel(df, path):
    # Excel 파일 작성
    wb = Workbook()
    ws = wb.active

    # 헤더 추가
    ws.append(df.columns.tolist())

    # 데이터 추가
    for index, row in df.iterrows():
        ws.append(row.tolist())

    # 셀 병합 및 너비 조정
    merge_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더
        folder_name = row[0]
        if folder_name:  # 폴더 이름이 있는 경우에만
            if folder_name not in merge_rows:
                merge_rows[folder_name] = []
            merge_rows[folder_name].append(index + 2)  # 데이터 행 인덱스 저장 (1 기반)

    # 같은 셀 이름 병합
    for folder_name, rows in merge_rows.items():
        if len(rows) > 1:  # 두 개 이상의 행이 있는 경우에만 병합
            ws.merge_cells(start_row=rows[0], start_column=1, end_row=rows[-1], end_column=1)

    # 셀 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # A, B, C, ...
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 너비 조정 (여유 공간 추가)
        ws.column_dimensions[column].width = adjusted_width

    # Excel 파일 저장
    wb.save(path)

initial_path  = r"D:\1_기술혁신팀\2_코드리뷰\2024\2024_권지웅_HMI_P4_HVAC"

try:
    result_df = list_files_and_folders(initial_path )
    print(result_df)

    # DataFrame을 Excel 파일로 저장
    # 현재 시간 추가하여 Excel 파일 이름 생성
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file_name = f"file_list_{current_time}.xlsx"
    # Excel 파일 저장 경로 설정 (초기 경로와 같은 경로에 저장)
    excel_file_path = os.path.join(initial_path, excel_file_name)

    save_to_excel(result_df, excel_file_path)
    print(f"DataFrame이 Excel 파일로 저장되었습니다: {excel_file_path}")

except ValueError as e:
    print(e)
